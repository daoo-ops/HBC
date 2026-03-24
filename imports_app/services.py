import csv
import io
import re
import unicodedata
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable, Union

import xlrd
from django.db import transaction
from openpyxl import load_workbook

from auditing.services import get_instance_snapshot, log_model_event
from clients.models import Client
from clients.utils import calculate_ruc_dv_from_base, extract_ruc_digit, normalize_ruc, normalize_ruc_base


SourceType = Union[Path, tuple[str, bytes]]


@dataclass
class PreviewResult:
    rows: list[dict]
    total_rows: int
    valid_rows: int
    duplicates_count: int
    missing_name_count: int
    missing_ruc_count: int
    warnings: list[str]


def _to_decimal(value) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    clean = str(value).strip().replace(".", "").replace(",", ".")
    try:
        return Decimal(clean)
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _read_source(source: SourceType) -> tuple[str, bytes]:
    if isinstance(source, Path):
        return source.name, source.read_bytes()
    name, payload = source
    return name, payload


def _canonical_name(value: str) -> str:
    data = (value or "").strip().upper()
    data = "".join(
        c for c in unicodedata.normalize("NFD", data) if unicodedata.category(c) != "Mn"
    )
    data = re.sub(r"\s+", " ", data)
    return data


def _zone_from_sheet(sheet_name: str):
    cleaned = _canonical_name(sheet_name)
    if "SANTA RITA" in cleaned:
        return Client.Zone.SANTA_RITA, Client.Status.ACTIVE
    if "KM 32" in cleaned:
        return Client.Zone.KM_32, Client.Status.ACTIVE
    if "MISION" in cleaned:
        return Client.Zone.COMISIONES, Client.Status.ACTIVE
    if "COMISION" in cleaned:
        return Client.Zone.COMISIONES, Client.Status.ACTIVE
    if "SUSPENSO" in cleaned or "SUSPEND" in cleaned:
        return Client.Zone.SUSPENDIDO, Client.Status.SUSPENDED
    if "KATUETE" in cleaned:
        return Client.Zone.KATUETE, Client.Status.ACTIVE
    return Client.Zone.OTHER, Client.Status.ACTIVE


HEADER_ALIASES = [
    ("ruc_base", {"RUC BASE", "RUC_BASE", "RUCBASE"}),
    ("dv", {"DV", "DIGITO VERIFICADOR", "DIGITO", "D V"}),
    ("ruc", {"RUC"}),
    ("name", {"NOMBRE", "CLIENTE", "SUCURSAL", "RAZON SOCIAL"}),
    ("value_gs", {"VALOR GS", "GS", "VALOR_GS", "MONTO GS", "GUARANIES", "GUARANI"}),
    ("value_usd", {"VALOR USD", "USD", "VALOR_USD", "MONTO USD", "DOLAR", "DOLARES"}),
    ("obligation", {"OBLIGACION", "OBLIGACIÓN", "TIPO PRESENTACION", "TIPO PRESENTACIÓN", "IMPUESTO"}),
    ("zone", {"ZONA", "UBICACION", "UBICACIÓN", "SUCURSAL / UBICACION", "SUCURSAL/UBICACION"}),
]


def _header_token(value) -> str:
    return _canonical_name(str(value or "")).replace("_", " ").strip()


def _field_from_header_token(token: str):
    for field_name, aliases in HEADER_ALIASES:
        for alias in aliases:
            if token == alias or alias in token:
                return field_name
    return None


def _detect_generic_header(ws, max_scan_rows: int = 12):
    best_row = None
    best_mapping = {}
    best_score = 0

    for row_idx in range(1, min(ws.max_row, max_scan_rows) + 1):
        mapping = {}
        for col_idx in range(1, ws.max_column + 1):
            token = _header_token(ws.cell(row_idx, col_idx).value)
            if not token:
                continue
            field_name = _field_from_header_token(token)
            if field_name and field_name not in mapping:
                mapping[field_name] = col_idx

        has_name = "name" in mapping
        has_ruc_data = "ruc" in mapping or "ruc_base" in mapping
        score = len(mapping)
        if has_name and has_ruc_data and score > best_score:
            best_row = row_idx
            best_mapping = mapping
            best_score = score

    if best_row is None:
        return None, {}
    return best_row, best_mapping


def _parse_generic_xlsx_sheet(ws, sheet_name: str) -> list[dict]:
    header_row, mapping = _detect_generic_header(ws)
    if header_row is None:
        return []

    zone_from_sheet, status_from_sheet = _zone_from_sheet(sheet_name)
    rows = []

    for row_idx in range(header_row + 1, ws.max_row + 1):
        name = str(ws.cell(row_idx, mapping["name"]).value or "").strip()
        raw_ruc = str(ws.cell(row_idx, mapping["ruc"]).value or "").strip() if "ruc" in mapping else ""
        raw_ruc_base = (
            str(ws.cell(row_idx, mapping["ruc_base"]).value or "").strip() if "ruc_base" in mapping else ""
        )
        raw_dv = str(ws.cell(row_idx, mapping["dv"]).value or "").strip() if "dv" in mapping else ""
        zone_hint = str(ws.cell(row_idx, mapping["zone"]).value or "").strip() if "zone" in mapping else ""

        ruc = normalize_ruc(raw_ruc)
        ruc_base = normalize_ruc_base(raw_ruc_base)
        dv = normalize_ruc(raw_dv).replace("-", "")

        if not ruc and ruc_base:
            ruc = f"{ruc_base}-{dv}" if dv else ruc_base
        elif ruc and "-" not in ruc and dv:
            base = normalize_ruc_base(ruc)
            if base:
                ruc = f"{base}-{dv}"

        if not name and not ruc:
            continue
        if name and name.strip().isdigit() and not ruc:
            continue
        if _canonical_name(name) in {"CLIENTE", "SUCURSAL", "N", "NO", "NRO", "Nº"}:
            continue

        zone, status = _zone_from_sheet(zone_hint) if zone_hint else (zone_from_sheet, status_from_sheet)

        rows.append(
            {
                "name": name,
                "ruc": ruc,
                "zone": zone,
                "status": status,
                "presentation_type": (
                    str(ws.cell(row_idx, mapping["obligation"]).value or "").strip() if "obligation" in mapping else ""
                ),
                "monthly_amount_pyg": (
                    _to_decimal(ws.cell(row_idx, mapping["value_gs"]).value) if "value_gs" in mapping else Decimal("0")
                ),
                "monthly_amount_usd": (
                    _to_decimal(ws.cell(row_idx, mapping["value_usd"]).value) if "value_usd" in mapping else Decimal("0")
                ),
            }
        )

    return rows


def _parse_master_xlsx(payload: bytes) -> list[dict]:
    workbook = load_workbook(io.BytesIO(payload), data_only=True)
    rows = []

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        if ws.max_row < 2:
            continue

        col_client = str(ws.cell(2, 2).value or "").strip().upper()
        if col_client == "CLIENTE":
            zone, default_status = _zone_from_sheet(sheet_name)

            for row_idx in range(4, ws.max_row + 1):
                name = str(ws.cell(row_idx, 2).value or "").strip()
                ruc = normalize_ruc(str(ws.cell(row_idx, 3).value or ""))
                amount_pyg = _to_decimal(ws.cell(row_idx, 4).value)
                amount_usd = _to_decimal(ws.cell(row_idx, 5).value)
                obligation = str(ws.cell(row_idx, 6).value or "").strip()

                if not name and not ruc:
                    continue
                if _canonical_name(name) in {"CLIENTE", "N", "NO", "NRO", "Nº"}:
                    continue
                if name and name.strip().isdigit() and not ruc:
                    continue

                rows.append(
                    {
                        "name": name,
                        "ruc": ruc,
                        "zone": zone,
                        "status": default_status,
                        "presentation_type": obligation,
                        "monthly_amount_pyg": amount_pyg,
                        "monthly_amount_usd": amount_usd,
                    }
                )
            continue

        rows.extend(_parse_generic_xlsx_sheet(ws, sheet_name))

    return rows


def _parse_master_csv(payload: bytes) -> list[dict]:
    decoded = payload.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(decoded))

    rows = []
    for row in reader:
        normalized = {(_canonical_name(k)): (v or "").strip() for k, v in row.items() if k}

        name = normalized.get("NOMBRE") or normalized.get("CLIENTE") or ""
        ruc = normalize_ruc(normalized.get("RUC", ""))
        zone_raw = normalized.get("ZONA", "")
        zone, _ = _zone_from_sheet(zone_raw)

        status_raw = _canonical_name(normalized.get("ESTADO", ""))
        status = Client.Status.SUSPENDED if "SUSP" in status_raw else Client.Status.ACTIVE

        rows.append(
            {
                "name": name,
                "ruc": ruc,
                "zone": zone,
                "status": status,
                "presentation_type": normalized.get("TIPO PRESENTACION", "")
                or normalized.get("OBLIGACION", ""),
                "monthly_amount_pyg": _to_decimal(normalized.get("MONTO MENSUAL", "") or normalized.get("VALOR GS", "")),
                "monthly_amount_usd": _to_decimal(normalized.get("VALOR USD", "")),
            }
        )

    return rows


def _parse_aux_xls(payload: bytes) -> dict[str, str]:
    workbook = xlrd.open_workbook(file_contents=payload)
    sheet = workbook.sheet_by_index(0)
    ruc_map = {}

    for row_idx in range(1, sheet.nrows):
        raw_ruc = normalize_ruc(str(sheet.cell_value(row_idx, 2)).strip())
        if not raw_ruc:
            continue
        base = normalize_ruc_base(raw_ruc)
        if base and "-" in raw_ruc:
            ruc_map[base] = raw_ruc
    return ruc_map


def _parse_aux_xlsx(payload: bytes) -> dict[str, str]:
    workbook = load_workbook(io.BytesIO(payload), data_only=True)
    ruc_map = {}

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        for row_idx in range(1, ws.max_row + 1):
            raw_ruc = normalize_ruc(str(ws.cell(row_idx, 3).value or "").strip())
            if not raw_ruc:
                continue
            base = normalize_ruc_base(raw_ruc)
            if base and "-" in raw_ruc:
                ruc_map[base] = raw_ruc

    return ruc_map


def _parse_aux_csv(payload: bytes) -> dict[str, str]:
    decoded = payload.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(decoded))
    ruc_map = {}
    for row in reader:
        ruc = normalize_ruc(row.get("RUC", ""))
        if not ruc:
            continue
        base = normalize_ruc_base(ruc)
        if base and "-" in ruc:
            ruc_map[base] = ruc
    return ruc_map


def load_master_records(master_source: SourceType) -> list[dict]:
    name, payload = _read_source(master_source)
    ext = Path(name).suffix.lower()
    if ext == ".xlsx":
        return _parse_master_xlsx(payload)
    if ext == ".csv":
        return _parse_master_csv(payload)
    raise ValueError(f"Formato de archivo maestro no soportado: {name}")


def load_aux_ruc_map(aux_sources: Iterable[SourceType]) -> dict[str, str]:
    merged = {}
    for source in aux_sources:
        name, payload = _read_source(source)
        ext = Path(name).suffix.lower()
        if ext == ".xls":
            merged.update(_parse_aux_xls(payload))
        elif ext == ".xlsx":
            merged.update(_parse_aux_xlsx(payload))
        elif ext == ".csv":
            merged.update(_parse_aux_csv(payload))
    return merged


def _deduplicate(rows: list[dict]) -> tuple[list[dict], int]:
    unique = []
    seen = set()
    duplicates = 0

    for row in rows:
        ruc_base = normalize_ruc_base(row.get("ruc", ""))
        if ruc_base:
            key = ("ruc", ruc_base)
        else:
            key = ("name_zone", _canonical_name(row.get("name", "")), row.get("zone"))

        if key in seen:
            duplicates += 1
            continue
        seen.add(key)
        unique.append(row)

    return unique, duplicates


def build_preview(master_source: SourceType, aux_sources: Iterable[SourceType]) -> PreviewResult:
    rows = load_master_records(master_source)
    aux_map = load_aux_ruc_map(aux_sources)

    missing_name = 0
    missing_ruc = 0

    for row in rows:
        row["name"] = (row.get("name") or "").strip()
        row["ruc"] = normalize_ruc(row.get("ruc", ""))
        if not row["name"]:
            missing_name += 1
        if not row["ruc"]:
            missing_ruc += 1

        base = normalize_ruc_base(row["ruc"])
        if base:
            has_dv = bool(extract_ruc_digit(row["ruc"]))
            if not has_dv:
                if base in aux_map:
                    row["ruc"] = aux_map[base]
                else:
                    dv = calculate_ruc_dv_from_base(base)
                    if dv:
                        row["ruc"] = f"{base}-{dv}"

    valid_rows = [r for r in rows if r.get("name")]
    deduped, duplicates = _deduplicate(valid_rows)

    warnings = []
    if missing_name:
        warnings.append(f"{missing_name} filas sin nombre fueron omitidas.")
    if missing_ruc:
        warnings.append(f"{missing_ruc} filas sin RUC detectadas.")
    if duplicates:
        warnings.append(f"{duplicates} filas duplicadas fueron descartadas en preview.")

    return PreviewResult(
        rows=deduped,
        total_rows=len(rows),
        valid_rows=len(deduped),
        duplicates_count=duplicates,
        missing_name_count=missing_name,
        missing_ruc_count=missing_ruc,
        warnings=warnings,
    )


@transaction.atomic
def commit_preview(preview: PreviewResult, actor):
    created = 0
    updated = 0

    for row in preview.rows:
        ruc = normalize_ruc(row.get("ruc", ""))
        ruc_base = normalize_ruc_base(ruc)
        ruc_dv = extract_ruc_digit(ruc)
        if ruc_base and not ruc_dv:
            guessed_dv = calculate_ruc_dv_from_base(ruc_base)
            if guessed_dv:
                ruc_dv = guessed_dv
                if "-" not in ruc:
                    ruc = f"{ruc_base}-{ruc_dv}"
        name = row.get("name", "")
        zone = row.get("zone") or Client.Zone.OTHER

        instance = None
        if ruc_base:
            instance = Client.objects.filter(ruc_base=ruc_base).order_by("id").first()
        if instance is None:
            instance = (
                Client.objects.filter(name__iexact=name.strip(), zone=zone)
                .order_by("id")
                .first()
            )

        payload = {
            "name": name,
            "ruc": ruc,
            "ruc_dv": ruc_dv,
            "zone": zone,
            "status": row.get("status", Client.Status.ACTIVE),
            "presentation_type": row.get("presentation_type", ""),
            "monthly_amount_pyg": row.get("monthly_amount_pyg", Decimal("0")),
            "monthly_amount_usd": row.get("monthly_amount_usd", Decimal("0")),
            "is_deleted": False,
            "deleted_at": None,
        }

        if instance is None:
            instance = Client.objects.create(**payload)
            created += 1
            log_model_event(
                actor=actor,
                action="import_create",
                instance=instance,
                after_data=get_instance_snapshot(instance),
                metadata={"source": "imports/clients/commit"},
            )
        else:
            before = get_instance_snapshot(instance)
            for key, value in payload.items():
                setattr(instance, key, value)
            instance.save()
            updated += 1
            log_model_event(
                actor=actor,
                action="import_update",
                instance=instance,
                before_data=before,
                after_data=get_instance_snapshot(instance),
                metadata={"source": "imports/clients/commit"},
            )

    return {
        "created": created,
        "updated": updated,
        "processed": preview.valid_rows,
    }
